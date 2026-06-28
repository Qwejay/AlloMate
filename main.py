import sys
import os
import shutil
import tempfile
import re
import time
import stat
import json
import locale
import math
import threading 
import csv
import random
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                               QPushButton, QLabel, QLineEdit, QTextEdit, QFileDialog, 
                               QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView, 
                               QStatusBar, QStackedWidget, QScrollArea, QFrame, QMessageBox,
                               QAbstractItemView, QGridLayout, QComboBox)
from PySide6.QtCore import Qt, Signal, QThread
from PySide6.QtGui import QFont, QColor, QBrush, QIcon

os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"
os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"

def get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def get_app_data_dir():
    if os.name == 'nt':
        base_dir = os.getenv('APPDATA') or os.path.expanduser('~/AppData/Roaming')
    else:
        base_dir = os.path.expanduser('~/.config')
    
    app_dir = os.path.join(base_dir, "AlloMate")
    if not os.path.exists(app_dir):
        try:
            os.makedirs(app_dir, exist_ok=True)
        except Exception:
            app_dir = tempfile.gettempdir()
    return app_dir

COLORS = {
    "primary": "#2563EB", 
    "primary_hover": "#1D4ED8", 
    "secondary": "#64748B",
    "bg_main": "#F1F5F9",       
    "bg_sidebar": "#0F172A",    
    "bg_card": "#FFFFFF",
    "text_dark": "#1E293B", 
    "text_light": "#F8FAFC", 
    "text_dim": "#94A3B8",
    "success": "#10B981", 
    "success_hover": "#059669",
    "danger": "#EF4444",
    "danger_hover": "#DC2626",
    "border": "#CBD5E1",
    "row_even": "#FFFFFF",      
    "row_odd": "#E2E8F0"        
}

FONTS = {
    "h1": ("Microsoft YaHei UI", 16, "bold"), 
    "h2": ("Microsoft YaHei UI", 11, "bold"),
    "body": ("Microsoft YaHei UI", 10), 
    "small": ("Microsoft YaHei UI", 9)
}

QSS_STYLE = f"""
QMainWindow {{
    background-color: {COLORS['bg_main']};
}}
QFrame#Sidebar {{
    background-color: {COLORS['bg_sidebar']};
}}
QFrame#Card {{
    background-color: {COLORS['bg_card']};
    border: 1px solid {COLORS['border']};
    border-radius: 8px;
}}
QLabel {{
    color: {COLORS['text_dark']};
    font-family: "Microsoft YaHei UI";
}}
QLineEdit, QTextEdit {{
    background-color: #FFFFFF;
    border: 1px solid {COLORS['border']};
    border-radius: 6px;
    padding: 6px 10px;
    color: {COLORS['text_dark']};
    font-family: "Microsoft YaHei UI";
}}
QLineEdit:focus, QTextEdit:focus {{
    border: 1px solid {COLORS['primary']};
}}
QPushButton#SidebarBtn {{
    color: {COLORS['text_light']};
    background-color: transparent;
    border: none;
    text-align: left;
    padding: 12px 25px;
    font-size: 14px;
    font-family: "Microsoft YaHei UI";
}}
QPushButton#SidebarBtn:hover {{
    background-color: #334155;
}}
QPushButton#SidebarBtn:checked {{
    background-color: {COLORS['primary']};
    font-weight: bold;
}}
QPushButton#PrimaryBtn {{
    background-color: {COLORS['primary']};
    color: white;
    border-radius: 6px;
    padding: 8px 18px;
    font-weight: bold;
    font-family: "Microsoft YaHei UI";
}}
QPushButton#PrimaryBtn:hover {{
    background-color: {COLORS['primary_hover']};
}}
QPushButton#SuccessBtn {{
    background-color: {COLORS['success']};
    color: white;
    border-radius: 6px;
    padding: 8px 18px;
    font-weight: bold;
    font-family: "Microsoft YaHei UI";
}}
QPushButton#SuccessBtn:hover {{
    background-color: {COLORS['success_hover']};
}}
QPushButton#DangerBtn {{
    background-color: {COLORS['danger']};
    color: white;
    border-radius: 6px;
    padding: 8px 18px;
    font-weight: bold;
    font-family: "Microsoft YaHei UI";
}}
QPushButton#DangerBtn:hover {{
    background-color: {COLORS['danger_hover']};
}}
QTableWidget {{
    gridline-color: {COLORS['border']};
    background-color: #FFFFFF;
    border: none;
    font-family: "Microsoft YaHei UI";
}}
QHeaderView::section {{
    background-color: #E2E8F0;
    color: {COLORS['text_dark']};
    padding: 8px;
    font-weight: bold;
    border: 1px solid {COLORS['border']};
    font-family: "Microsoft YaHei UI";
}}
QTabWidget::pane {{
    border: 1px solid {COLORS['border']};
    background-color: #FFFFFF;
    border-radius: 6px;
}}
QTabBar::tab {{
    background-color: #E2E8F0;
    border: 1px solid {COLORS['border']};
    padding: 8px 20px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    font-family: "Microsoft YaHei UI";
}}
QTabBar::tab:selected {{
    background-color: #FFFFFF;
    border-bottom-color: transparent;
}}
QStatusBar {{
    background-color: #FFFFFF;
    border-top: 1px solid {COLORS['border']};
    font-family: "Microsoft YaHei UI";
}}
QComboBox {{
    background-color: #FFFFFF;
    border: 1px solid {COLORS['primary']};
    border-radius: 4px;
    padding: 2px 10px;
    color: {COLORS['text_dark']};
    font-family: "Microsoft YaHei UI";
}}
"""

class AllocationModel:
    def __init__(self):
        self.user_data_dir = get_app_data_dir()
        self.config_file = os.path.join(self.user_data_dir, "AlloMate.json")
        self.history_file = os.path.join(self.user_data_dir, "AlloMate_history.json")
        self.config = self.load_config()
        self.history = self.load_history()

    def get_default_config(self):
        return {
            "object_type": "姓名",
            "object_names": [f"小朋友{i}" for i in range(1, 31)],
            "group_type": "区域",
            "group_names": ["娃娃家", "建构区", "语言区", "科学区", "美工区"],
            "cycle_type": "第",
            "cycle_count": 20,
            "min_visits": 2, 
            "max_visits": 5,           
            "max_obj_allocations": 15, 
            "group_capacity": {
                "娃娃家": 3, "建构区": 6, "语言区": 8, "科学区": 10, "美工区": 8
            },
            "group_weights": {         
                "娃娃家": 1, "建构区": 1, "语言区": 1, "科学区": 1, "美工区": 1
            },
            "priority_rules": [],
            "checked_symbol": "✅"
        }

    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    default = self.get_default_config()
                    for k, v in default.items():
                        if k not in data: data[k] = v
                    return data
            except: pass
        default = self.get_default_config()
        self.save_config(default)
        return default

    def save_config(self, new_config):
        self.config = new_config
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(new_config, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            print(f"Error saving config: {e}")
            return False

    def load_history(self):
        if not os.path.exists(self.history_file): return []
        try:
            with open(self.history_file, 'r', encoding='utf-8') as f: return json.load(f)
        except: return []

    def save_history(self, allocation):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.history.insert(0, {"timestamp": timestamp, "allocation": allocation})
        self.history = self.history[:10]
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, ensure_ascii=False, indent=4)
        except: pass

    def generate_allocation(self):
        cfg = self.config
        obj_names = cfg['object_names']
        grp_names = cfg['group_names']
        grp_cap = cfg['group_capacity']
        grp_weight = cfg.get('group_weights', {g: 1 for g in grp_names})
        cycles = list(range(1, cfg['cycle_count'] + 1))
        
        min_visits = cfg.get('min_visits', 0)
        max_visits = cfg.get('max_visits', 99)
        max_obj_alloc = cfg.get('max_obj_allocations', 99)
        
        allocation = {obj: {g: [] for g in grp_names} for obj in obj_names}
        pri_rules = {}
        for r in cfg.get('priority_rules', []):
            if ':' in r:
                p = r.split(':')
                if len(p) >= 2: pri_rules[p[0].strip()] = p[1].strip()

        def weighted_choice(choices, weights_dict):
            if not choices: return None
            w = [weights_dict.get(c, 1) for c in choices]
            if sum(w) == 0:
                return random.choice(choices)
            return random.choices(choices, weights=w, k=1)[0]

        for cycle in cycles:
            grp_count = {g: 0 for g in grp_names}
            shuffled = list(obj_names)
            random.shuffle(shuffled)
            for obj in shuffled:
                current_total_allocs = sum(len(allocation[obj][g]) for g in grp_names)
                if current_total_allocs >= max_obj_alloc:
                    continue 
                
                avail = [g for g in grp_names if grp_count[g] < grp_cap.get(g, 0)]
                avail = [g for g in avail if len(allocation[obj][g]) < max_visits]
                
                if not avail: continue
                
                last_grp = None
                for g, c_list in allocation[obj].items():
                    if cycle - 1 in c_list:
                        last_grp = g
                        break
                
                pref = pri_rules.get(obj)
                if pref in avail: 
                    chosen = pref
                else:
                    deficit_groups = [g for g in avail if len(allocation[obj][g]) < min_visits]
                    
                    if deficit_groups:
                        filtered_deficit = [g for g in deficit_groups if g != last_grp]
                        chosen = weighted_choice(filtered_deficit if filtered_deficit else deficit_groups, grp_weight)
                    else:
                        if last_grp and last_grp in avail and len(avail) > 1:
                            filtered = [g for g in avail if g != last_grp]
                            chosen = weighted_choice(filtered if filtered else avail, grp_weight)
                        else: 
                            chosen = weighted_choice(avail, grp_weight)
                            
                allocation[obj][chosen].append(cycle)
                grp_count[chosen] += 1
        self.save_history(allocation)
        return allocation

    def allocation_to_data(self, allocation):
        cycles = list(range(1, self.config['cycle_count'] + 1))
        result = {}
        for group in self.config['group_names']:
            headers = [self.config['object_type']] + [f"{self.config['cycle_type']}{c}" for c in cycles]
            rows = []
            for obj in self.config['object_names']:
                row = [obj]
                for cycle in cycles:
                    if cycle in allocation.get(obj, {}).get(group, []):
                        row.append(self.config['checked_symbol'])
                    else:
                        row.append('')
                rows.append(row)
            result[group] = {"headers": headers, "rows": rows}
        return result

    def export_data(self, allocation, file_path):
        data_map = self.allocation_to_data(allocation)
        try:
            if file_path.endswith('.csv'):
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    cycles = list(range(1, self.config['cycle_count'] + 1))
                    headers = [self.config['object_type']] + [f"{self.config['cycle_type']}{c}" for c in cycles]
                    writer.writerow(headers)
                    for obj in self.config['object_names']:
                        row = [obj]
                        for cycle in cycles:
                            assigned_group = ""
                            for group in self.config['group_names']:
                                if cycle in allocation.get(obj, {}).get(group, []):
                                    assigned_group = group
                                    break
                            row.append(assigned_group)
                        writer.writerow(row)
            else:
                wb = Workbook()
                thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                ws_overview = wb.active
                ws_overview.title = "全景分配总览"
                cycles = list(range(1, self.config['cycle_count'] + 1))
                headers_overview = [self.config['object_type']] + [f"{self.config['cycle_type']}{c}" for c in cycles]
                ws_overview.append(headers_overview)
                
                for obj in self.config['object_names']:
                    row = [obj]
                    for cycle in cycles:
                        assigned_group = ""
                        for group in self.config['group_names']:
                            if cycle in allocation.get(obj, {}).get(group, []):
                                assigned_group = group
                                break
                        row.append(assigned_group)
                    ws_overview.append(row)
                
                for row in ws_overview.iter_rows():
                    for cell in row: cell.border = thin
                
                for group, data in data_map.items():
                    ws = wb.create_sheet(group)
                    ws.append(data['headers'])
                    for row in data['rows']: ws.append(row)
                    for row in ws.iter_rows():
                        for cell in row: cell.border = thin
                wb.save(file_path)
            return True, "✅ 数据已成功导出至：" + os.path.basename(file_path)
        except Exception as e:
            return False, "❌ 导出失败：" + str(e)


class ExportThread(QThread):
    finished_signal = Signal(bool, str)

    def __init__(self, model, allocation, file_path):
        super().__init__()
        self.model = model
        self.allocation = allocation
        self.file_path = file_path

    def run(self):
        success, msg = self.model.export_data(self.allocation, self.file_path)
        self.finished_signal.emit(success, msg)


class ResultsPage(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.current_allocation = None
        self.table_widgets = [] 
        self.active_edit_cell = None 
        self.init_ui()

    def create_card_layout(self):
        f = QFrame()
        f.setObjectName("Card")
        return f

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 30)
        layout.setSpacing(20)

        toolbar = self.create_card_layout()
        tb_layout = QHBoxLayout(toolbar)
        tb_layout.setContentsMargins(20, 15, 20, 15)

        lbl_title = QLabel("📅 分配结果")
        lbl_title.setStyleSheet("font-size: 20px; font-weight: bold;")
        tb_layout.addWidget(lbl_title)

        lbl_tip = QLabel("💡 在【全景总览】中双击单元格，可直接下拉框置换区域")
        lbl_tip.setStyleSheet(f"color: {COLORS['secondary']}; font-size: 12px;")
        tb_layout.addWidget(lbl_tip)
        tb_layout.addStretch()

        self.btn_refresh = QPushButton("🔄 重新生成")
        self.btn_refresh.setObjectName("PrimaryBtn")
        self.btn_refresh.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_refresh.clicked.connect(self.refresh_data)
        tb_layout.addWidget(self.btn_refresh)

        self.btn_export = QPushButton("📥 导出 Excel")
        self.btn_export.setObjectName("SuccessBtn")
        self.btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_export.clicked.connect(self.export_data)
        tb_layout.addWidget(self.btn_export)

        layout.addWidget(toolbar)

        # 选项卡
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

    def show_allocation(self, allocation):
        self.current_allocation = allocation
        self.tabs.clear()
        self.table_widgets.clear()
        self.active_edit_cell = None 

        if not allocation: return

        cycles = list(range(1, self.app.model.config['cycle_count'] + 1))

        tab_overview = QWidget()
        overview_layout = QVBoxLayout(tab_overview)
        overview_layout.setContentsMargins(0, 0, 0, 0)

        headers_overview = [self.app.model.config['object_type']] + [f"{self.app.model.config['cycle_type']}{c}" for c in cycles]
        rows_overview = []

        for obj in self.app.model.config['object_names']:
            row = [obj]
            for cycle in cycles:
                assigned_group = "轮空" 
                for group in self.app.model.config['group_names']:
                    if cycle in allocation.get(obj, {}).get(group, []):
                        assigned_group = group
                        break
                row.append(assigned_group)
            rows_overview.append(row)

        self.overview_table = QTableWidget()
        self.overview_table.setColumnCount(len(headers_overview))
        self.overview_table.setRowCount(len(rows_overview))
        self.overview_table.setHorizontalHeaderLabels(headers_overview)
        self.overview_table.horizontalHeader().setDefaultSectionSize(100)
        self.overview_table.verticalHeader().setVisible(False)
        self.overview_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.overview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)

        self.overview_table.blockSignals(True)
        for r_idx, row in enumerate(rows_overview):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if c_idx == 0:
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable) # 姓名列锁定
                else:
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEditable)
                
                bg_color = QColor(COLORS['row_even']) if r_idx % 2 == 0 else QColor(COLORS['row_odd'])
                item.setBackground(QBrush(bg_color))
                self.overview_table.setItem(r_idx, c_idx, item)
        self.overview_table.blockSignals(False)

        self.overview_table.cellDoubleClicked.connect(self.on_overview_cell_double_clicked)
        self.overview_table.cellClicked.connect(self.on_overview_cell_clicked)
        
        overview_layout.addWidget(self.overview_table)
        self.tabs.addTab(tab_overview, "  📊 总览 (双击修改)  ")

        data_map = self.app.model.allocation_to_data(allocation)

        for group_name, data in data_map.items():
            tab_widget = QWidget()
            tab_layout = QVBoxLayout(tab_widget)
            tab_layout.setContentsMargins(0, 0, 0, 0)

            headers = data['headers']
            rows = data['rows']

            table = QTableWidget()
            table.setColumnCount(len(headers))
            table.setRowCount(len(rows))
            table.setHorizontalHeaderLabels(headers)
            table.horizontalHeader().setDefaultSectionSize(80)
            table.verticalHeader().setVisible(False)
            table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
            table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)

            table.blockSignals(True)
            for row_idx, row in enumerate(rows):
                for col_idx, val in enumerate(row):
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable) # 归档卡片全部只读，防止改错

                    bg_color = QColor(COLORS['row_even']) if row_idx % 2 == 0 else QColor(COLORS['row_odd'])
                    item.setBackground(QBrush(bg_color))
                    table.setItem(row_idx, col_idx, item)
            table.blockSignals(False)

            tab_layout.addWidget(table)
            self.tabs.addTab(tab_widget, f" {group_name} ")

        self.app.set_status(f"📅 分配方案已生成 (生成时间: {datetime.now().strftime('%H:%M:%S')})")

    def on_overview_cell_clicked(self, row, col):
        if self.active_edit_cell:
            act_r, act_c = self.active_edit_cell
            if act_r != row or act_c != col:
                self.commit_active_editor()

    def commit_active_editor(self):
        if self.active_edit_cell:
            r, c = self.active_edit_cell
            combo = self.overview_table.cellWidget(r, c)
            if combo:
                try:
                    combo.currentIndexChanged.disconnect()
                except Exception:
                    pass
                val = combo.currentText()
                self.overview_table.removeCellWidget(r, c)
                
                item = self.overview_table.item(r, c)
                if item:
                    item.setText(val)
                self.save_overview_data(r, c, val)
                
            self.active_edit_cell = None

    def on_overview_cell_double_clicked(self, row, col):
        if col == 0:
            self.app.set_status("⚠️ 权限限制：【姓名】列为系统核心索引，禁止在此修改！请前往全局设置调整名单。", is_error=True)
            return

        self.commit_active_editor()

        self.active_edit_cell = (row, col)

        combo = QComboBox()
        groups_list = ["轮空"] + self.app.model.config['group_names']
        combo.addItems(groups_list)

        current_val = self.overview_table.item(row, col).text()
        if current_val in groups_list:
            combo.setCurrentText(current_val)
        else:
            combo.setCurrentText("轮空")

        self.overview_table.setCellWidget(row, col, combo)
        combo.showPopup()

        combo.currentIndexChanged.connect(lambda idx, r=row, c=col, cb=combo: self.on_combo_index_changed(r, c, cb))

    def on_combo_index_changed(self, row, col, combo):
        val = combo.currentText()
        try:
            combo.currentIndexChanged.disconnect()
        except Exception:
            pass
        self.overview_table.removeCellWidget(row, col)
        
        item = self.overview_table.item(row, col)
        if item:
            item.setText(val)
            
        self.active_edit_cell = None
        self.save_overview_data(row, col, val)

    def save_overview_data(self, row, col, val):
        obj_name = self.overview_table.item(row, 0).text()
        cycle_num = col

        if obj_name in self.current_allocation:
            for g in self.app.model.config['group_names']:
                if cycle_num in self.current_allocation[obj_name].get(g, []):
                    self.current_allocation[obj_name][g].remove(cycle_num)
            
            if val != "轮空":
                if val not in self.current_allocation[obj_name]:
                    self.current_allocation[obj_name][val] = []
                self.current_allocation[obj_name][val].append(cycle_num)
                self.current_allocation[obj_name][val].sort()

            self.app.model.save_history(self.current_allocation)
            self.silent_refresh_tabs()
            self.app.set_status(f"✅ 在线置换成功：已将【{obj_name}】第 {cycle_num} 次的分配区域变更为【{val}】")

    def silent_refresh_tabs(self):
        current_idx = self.tabs.currentIndex()
        self.show_allocation(self.current_allocation)
        self.tabs.setCurrentIndex(current_idx)

    def refresh_data(self):
        alloc = self.app.model.generate_allocation()
        self.show_allocation(alloc)
        self.app.update_history_sidebar()

    def export_data(self):
        if not self.current_allocation: return
        
        path, _ = QFileDialog.getSaveFileName(self, "导出分配方案", "随机分配.xlsx", "Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not path: return

        self.app.set_status("⌛ 正在执行线程化 Excel 导出，请稍候...", duration=0)
        self.btn_export.setEnabled(False)
        self.btn_refresh.setEnabled(False)

        self.export_thread = ExportThread(self.app.model, self.current_allocation, path)
        self.export_thread.finished_signal.connect(self.on_export_finished)
        self.export_thread.start()

    def on_export_finished(self, success, msg):
        self.btn_export.setEnabled(True)
        self.btn_refresh.setEnabled(True)
        if success:
            self.app.set_status(msg, is_error=False, duration=6000)
        else:
            self.app.set_status(msg, is_error=True, duration=8000)


class ConfigPage(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.entries = {}
        self.text_widgets = {}
        self.init_ui()

    def create_card_layout(self):
        f = QFrame()
        f.setObjectName("Card")
        return f

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 30)
        layout.setSpacing(20)

        header = QFrame()
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(0, 0, 0, 0)
        
        lbl_title = QLabel("⚙️ 全局设置")
        lbl_title.setStyleSheet("font-size: 20px; font-weight: bold;")
        h_layout.addWidget(lbl_title)
        h_layout.addStretch()

        btn_save = QPushButton("💾 保存配置")
        btn_save.setObjectName("PrimaryBtn")
        btn_save.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_save.clicked.connect(self.save_settings)
        h_layout.addWidget(btn_save)
        layout.addWidget(header)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("background-color: transparent;")
        
        scroll_content = QWidget()
        scroll_content.setStyleSheet(f"background-color: {COLORS['bg_main']};")
        self.scroll_layout = QVBoxLayout(scroll_content)
        self.scroll_layout.setContentsMargins(0, 0, 0, 0)
        self.scroll_layout.setSpacing(20)

        card_obj = self.create_card_layout()
        card_obj_lay = QVBoxLayout(card_obj)
        card_obj_lay.setContentsMargins(25, 20, 25, 20)
        card_obj_lay.setSpacing(15)
        
        card_obj_lay.addWidget(QLabel("👥 对象配置", font=QFont("Microsoft YaHei UI", 11, QFont.Weight.Bold)))
        
        lay_row1 = QHBoxLayout()
        lay_row1.addWidget(QLabel("对象类型名称:"))
        self.entry_obj_type = QLineEdit()
        self.entries["object_type"] = self.entry_obj_type
        lay_row1.addWidget(self.entry_obj_type)
        lay_row1.addStretch()
        card_obj_lay.addLayout(lay_row1)

        lay_name_hdr = QHBoxLayout()
        lay_name_hdr.addWidget(QLabel("名单列表 (每行一个人名):"))
        lay_name_hdr.addStretch()
        
        btn_import = QPushButton("📂 导入名单文件")
        btn_import.setStyleSheet(f"color: {COLORS['primary']}; font-weight: bold; border: none; background: transparent;")
        btn_import.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_import.clicked.connect(self.import_names_from_file)
        lay_name_hdr.addWidget(btn_import)

        btn_demo = QPushButton("🧙 载入演示数据")
        btn_demo.setStyleSheet(f"color: {COLORS['secondary']}; font-weight: bold; border: none; background: transparent;")
        btn_demo.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_demo.clicked.connect(self.load_demo_data)
        lay_name_hdr.addWidget(btn_demo)
        card_obj_lay.addLayout(lay_name_hdr)

        self.text_obj_names = QTextEdit()
        self.text_obj_names.setFixedHeight(180)
        self.text_widgets["object_names"] = self.text_obj_names
        card_obj_lay.addWidget(self.text_obj_names)
        self.scroll_layout.addWidget(card_obj)

        card_cycle = self.create_card_layout()
        card_cycle_lay = QVBoxLayout(card_cycle)
        card_cycle_lay.setContentsMargins(25, 20, 25, 20)
        card_cycle_lay.setSpacing(15)
        card_cycle_lay.addWidget(QLabel("📅 周期配置与核心约束", font=QFont("Microsoft YaHei UI", 11, QFont.Weight.Bold)))

        lay_cycle_inputs = QGridLayout()
        lay_cycle_inputs.addWidget(QLabel("周期单位 (例如: 周/第/天):"), 0, 0)
        self.entry_cycle_type = QLineEdit()
        self.entries["cycle_type"] = self.entry_cycle_type
        lay_cycle_inputs.addWidget(self.entry_cycle_type, 0, 1)

        lay_cycle_inputs.addWidget(QLabel("总分配周期数 (数字):"), 0, 2)
        self.entry_cycle_count = QLineEdit()
        self.entries["cycle_count"] = self.entry_cycle_count
        lay_cycle_inputs.addWidget(self.entry_cycle_count, 0, 3)

        lay_cycle_inputs.addWidget(QLabel("各区域最少进入次数 (X):"), 1, 0)
        self.entry_min_v = QLineEdit()
        self.entries["min_visits"] = self.entry_min_v
        lay_cycle_inputs.addWidget(self.entry_min_v, 1, 1)

        lay_cycle_inputs.addWidget(QLabel("各区域最多进入次数 (Y):"), 1, 2)
        self.entry_max_v = QLineEdit()
        self.entries["max_visits"] = self.entry_max_v
        lay_cycle_inputs.addWidget(self.entry_max_v, 1, 3)

        lay_cycle_inputs.addWidget(QLabel("单人最高出勤总数限制:"), 2, 0)
        self.entry_max_allocs = QLineEdit()
        self.entries["max_obj_allocations"] = self.entry_max_allocs
        lay_cycle_inputs.addWidget(self.entry_max_allocs, 2, 1)
        card_cycle_lay.addLayout(lay_cycle_inputs)
        self.scroll_layout.addWidget(card_cycle)

        card_adv = self.create_card_layout()
        card_adv_lay = QVBoxLayout(card_adv)
        card_adv_lay.setContentsMargins(25, 20, 25, 20)
        card_adv_lay.setSpacing(15)
        card_adv_lay.addWidget(QLabel("🛡️ 高级约束规则", font=QFont("Microsoft YaHei UI", 11, QFont.Weight.Bold)))

        card_adv_lay.addWidget(QLabel("强制绑定分配规则 (格式: 姓名:区域名, 每行一项):"))
        self.text_rules = QTextEdit()
        self.text_rules.setFixedHeight(120)
        self.text_widgets["priority_rules"] = self.text_rules
        card_adv_lay.addWidget(self.text_rules)

        lay_symbol = QHBoxLayout()
        lay_symbol.addWidget(QLabel("系统已分配勾选标记 (如 ✅):"))
        self.entry_symbol = QLineEdit()
        self.entries["checked_symbol"] = self.entry_symbol
        lay_symbol.addWidget(self.entry_symbol)
        lay_symbol.addStretch()
        card_adv_lay.addLayout(lay_symbol)
        self.scroll_layout.addWidget(card_adv)

        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

    def import_names_from_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入名单文件", "", "Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not path: return
        
        names = []
        try:
            if path.endswith('.csv'):
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if row: names.append(row[0].strip())
            else:
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if row and row[0]:
                        names.append(str(row[0]).strip())
            
            if names and (names[0] == "姓名" or names[0] == "名单" or names[0] == "学生" or names[0] == self.entry_obj_type.text().strip()):
                names = names[1:]
            
            names = [n for n in names if n]
            if names:
                self.text_obj_names.setPlainText("\n".join(names))
                self.app.set_status(f"📂 成功从文件智能提取并导入了 {len(names)} 个名单数据！")
            else:
                self.app.set_status("⚠️ 导入失败：未在文件第一列检索到有效的人名名单", is_error=True)
        except Exception as e:
            self.app.set_status(f"❌ 导入出错：{e}", is_error=True)

    def load_demo_data(self):
        names = ["张伟", "王伟", "李娜", "刘洋", "陈静", "杨兵", "赵勇", "黄燕", "周涛", "吴丽",
                 "徐杰", "孙超", "胡军", "朱婷", "高飞", "林晨", "何伟", "郭强", "马莉", "罗荣",
                 "梁栋", "宋涛", "韩梅", "邓波", "冯钢", "曹宇", "彭程", "曾强", "肖芳", "田野"]
        
        self.entry_obj_type.setText("学生姓名")
        self.entry_cycle_type.setText("周")
        self.entry_cycle_count.setText("15")
        self.entry_min_v.setText("2")
        self.entry_max_v.setText("5")
        self.entry_max_allocs.setText("12")
        self.entry_symbol.setText("✅")
        
        self.text_obj_names.setPlainText("\n".join(names))
        self.text_rules.setPlainText("张伟:娃娃家\n李娜:科学区\n刘洋:美工区")
        self.app.set_status("🧙 AlloMate 30人标准分配教学演示模板已载入本地工作区")

    def load_values(self):
        cfg = self.app.model.config
        for key, widget in self.entries.items():
            widget.setText(str(cfg.get(key, "")))
        for key, widget in self.text_widgets.items():
            val = cfg.get(key, [])
            if isinstance(val, list):
                widget.setPlainText("\n".join(val))
            else:
                widget.setPlainText(str(val))

    def save_settings(self):
        new_config = self.app.model.config.copy()
        try:
            for key, widget in self.entries.items():
                val = widget.text().strip()
                if key in ["cycle_count", "min_visits", "max_visits", "max_obj_allocations"]:
                    val = int(val)
                new_config[key] = val
            for key, widget in self.text_widgets.items():
                content = widget.toPlainText().strip()
                lines = [line.strip() for line in content.split('\n') if line.strip()]
                new_config[key] = lines

            min_v = int(new_config.get("min_visits", 0))
            max_v = int(new_config.get("max_visits", 99))
            max_obj_alloc = int(new_config.get("max_obj_allocations", 99))
            cycle_c = int(new_config.get("cycle_count", 0))
            grp_len = len(new_config.get("group_names", []))

            if min_v > max_v:
                self.app.set_status("❌ 保存失败：最少进入次数(X) 不能大于 最多进入次数(Y)！", is_error=True, duration=8000)
                return
            if min_v * grp_len > max_obj_alloc:
                self.app.set_status(f"❌ 保存失败：最少次数保底总需求 ({min_v}×{grp_len}={min_v * grp_len}) 已超出单人出勤分配总量上限 ({max_obj_alloc})！", is_error=True, duration=8000)
                return
            if max_obj_alloc > cycle_c:
                self.app.set_status(f"❌ 保存失败：单人出勤分配总上限 ({max_obj_alloc}) 不能大于 周期总数 ({cycle_c})！", is_error=True, duration=8000)
                return

            if self.app.model.save_config(new_config):
                self.app.set_status("✅ 全局分配策略及约束指标保存成功", is_error=False)
                self.app.switch_to("results")
                self.app.pages["results"].refresh_data()
        except ValueError:
            self.app.set_status("⚠️ 保存失败：分配周期数量和最少次数限制必须是合法的纯数字格式！", is_error=True)


class GroupsPage(QWidget):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.init_ui()

    def create_card_layout(self):
        f = QFrame()
        f.setObjectName("Card")
        return f

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 30)
        layout.setSpacing(20)

        header = QFrame()
        h_layout = QHBoxLayout(header)
        h_layout.setContentsMargins(0, 0, 0, 0)

        lbl_title = QLabel("👥 分组管理")
        lbl_title.setStyleSheet("font-size: 20px; font-weight: bold;")
        h_layout.addWidget(lbl_title)
        h_layout.addStretch()

        btn_save = QPushButton("💾 保存分组")
        btn_save.setObjectName("PrimaryBtn")
        btn_save.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_save.clicked.connect(self.save_groups)
        h_layout.addWidget(btn_save)
        layout.addWidget(header)

        container = self.create_card_layout()
        c_layout = QVBoxLayout(container)
        c_layout.setContentsMargins(25, 25, 25, 25)
        c_layout.setSpacing(15)

        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("分组名称:"))
        self.name_entry = QLineEdit()
        self.name_entry.setFixedWidth(140)
        input_layout.addWidget(self.name_entry)

        input_layout.addWidget(QLabel("容量/班额:"))
        self.cap_entry = QLineEdit()
        self.cap_entry.setFixedWidth(60)
        self.cap_entry.setText("1")
        input_layout.addWidget(self.cap_entry)

        input_layout.addWidget(QLabel("权重几率:"))
        self.weight_entry = QLineEdit()
        self.weight_entry.setFixedWidth(60)
        self.weight_entry.setText("1")
        input_layout.addWidget(self.weight_entry)

        btn_add = QPushButton("➕ 添加")
        btn_add.setObjectName("SuccessBtn")
        btn_add.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_add.clicked.connect(self.add_group)
        input_layout.addWidget(btn_add)

        btn_update = QPushButton("✍️ 修改选中")
        btn_update.setObjectName("PrimaryBtn")
        btn_update.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_update.clicked.connect(self.update_group)
        input_layout.addWidget(btn_update)
        input_layout.addStretch()

        btn_delete = QPushButton("🗑️ 删除选中")
        btn_delete.setObjectName("DangerBtn")
        btn_delete.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_delete.clicked.connect(self.delete_selected)
        input_layout.addWidget(btn_delete)
        c_layout.addLayout(input_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["分组名称", "容量/最大班额", "分配权重 (权重越高被选几率越大)"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.itemSelectionChanged.connect(self.on_row_selected)
        c_layout.addWidget(self.table)

        layout.addWidget(container)

    def load_data(self):
        self.table.blockSignals(True)
        self.table.setRowCount(0)
        
        group_names = self.app.model.config.get('group_names', [])
        self.table.setRowCount(len(group_names))

        for row, name in enumerate(group_names):
            cap = self.app.model.config.get('group_capacity', {}).get(name, 1)
            weight = self.app.model.config.get('group_weights', {}).get(name, 1)

            item_name = QTableWidgetItem(name)
            item_name.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_name.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(row, 0, item_name)

            item_cap = QTableWidgetItem(str(cap))
            item_cap.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_cap.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(row, 1, item_cap)

            item_weight = QTableWidgetItem(str(weight))
            item_weight.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_weight.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(row, 2, item_weight)

        self.table.blockSignals(False)

    def on_row_selected(self):
        selected_rows = self.table.selectedItems()
        if not selected_rows: return
        row = selected_rows[0].row()
        self.name_entry.setText(self.table.item(row, 0).text())
        self.cap_entry.setText(self.table.item(row, 1).text())
        self.weight_entry.setText(self.table.item(row, 2).text())

    def add_group(self):
        name = self.name_entry.text().strip()
        cap = self.cap_entry.text().strip()
        weight = self.weight_entry.text().strip()

        if not name or not cap.isdigit() or not weight.isdigit():
            self.app.set_status("⚠️ 添加失败：请输入合法的分组名称、容量及分配权重数值！", is_error=True)
            return

        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).text() == name:
                self.app.set_status("⚠️ 添加失败：该分组已经存在于列表中，请勿重复添加", is_error=True)
                return

        row = self.table.rowCount()
        
        self.table.blockSignals(True)
        self.table.insertRow(row)
        
        item_name = QTableWidgetItem(name)
        item_name.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        item_name.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
        self.table.setItem(row, 0, item_name)

        item_cap = QTableWidgetItem(cap)
        item_cap.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        item_cap.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
        self.table.setItem(row, 1, item_cap)

        item_weight = QTableWidgetItem(weight)
        item_weight.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        item_weight.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
        self.table.setItem(row, 2, item_weight)
        
        self.table.blockSignals(False)

        self.name_entry.clear()
        self.app.set_status(f"➕ 已成功将临时分组【{name}】加入缓冲队列")

    def update_group(self):
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            self.app.set_status("⚠️ 修改失败：请先在下方列表中选择一整行要编辑的分组", is_error=True)
            return

        name = self.name_entry.text().strip()
        cap = self.cap_entry.text().strip()
        weight = self.weight_entry.text().strip()

        if not name or not cap.isdigit() or not weight.isdigit():
            self.app.set_status("⚠️ 修改失败：请输入合法的分组名称、容量及分配权重！", is_error=True)
            return

        row = selected_ranges[0].topRow()
        for r in range(self.table.rowCount()):
            if r != row and self.table.item(r, 0).text() == name:
                self.app.set_status("⚠️ 修改失败：修改后的分组名称与其他分组冲突！", is_error=True)
                return

        self.table.blockSignals(True)
        old_name = self.table.item(row, 0).text()
        self.table.item(row, 0).setText(name)
        self.table.item(row, 1).setText(cap)
        self.table.item(row, 2).setText(weight)
        self.table.blockSignals(False)
        self.app.set_status(f"✍️ 已成功将分组【{old_name}】属性同步更新为 【{name}】(容量: {cap}, 权重: {weight})")

    def delete_selected(self):
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            self.app.set_status("⚠️ 请选择要删除的分组数据行", is_error=True)
            return
        row = selected_ranges[0].topRow()
        name = self.table.item(row, 0).text()
        
        self.table.blockSignals(True)
        self.table.removeRow(row)
        self.table.blockSignals(False)
        
        self.app.set_status(f"🗑️ 已成功从缓冲队列移除分组：{name}")

    def save_groups(self):
        names, caps, weights = [], {}, {}
        for row in range(self.table.rowCount()):
            name = self.table.item(row, 0).text()
            cap = int(self.table.item(row, 1).text())
            weight = int(self.table.item(row, 2).text())
            names.append(name)
            caps[name] = cap
            weights[name] = weight

        if not names:
            self.app.set_status("❌ 保存失败：分配分组设置中必须至少包含一个有效分组！", is_error=True)
            return

        cfg = self.app.model.config
        cfg['group_names'] = names
        cfg['group_capacity'] = caps
        cfg['group_weights'] = weights

        if self.app.model.save_config(cfg):
            self.app.set_status("✅ 分配分组数据及权重指标同步更新保存成功", is_error=False)
            self.app.switch_to("results")
            self.app.pages["results"].refresh_data()


class Application(QMainWindow):
    def __init__(self):
        super().__init__()
        self.title_text = "AlloMate - 随机分配助手"
        self.setWindowTitle(self.title_text)
        self._status_timer = None

        icon_path = get_resource_path("logo.svg")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.adjust_geometry_to_resolution()
        self.setStyleSheet(QSS_STYLE)

        self.model = AllocationModel()
        self.init_ui()
        self.init_app_state()

    def adjust_geometry_to_resolution(self):
        screen = QApplication.primaryScreen().geometry()
        screen_width = screen.width()
        screen_height = screen.height()

        width = int(screen_width * 0.75)
        height = int(screen_height * 0.80)

        width = max(1024, min(width, 1366))
        height = max(700, min(height, 880))

        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        self.setGeometry(x, y, width, height)
        self.setMinimumSize(1024, 700)

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.sidebar = QFrame()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setFixedWidth(260)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(5)

        lbl_logo = QLabel("🤖 随机分配助手")
        lbl_logo.setStyleSheet("color: white; font-size: 16px; font-weight: bold; padding: 40px 25px;")
        lbl_logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sidebar_layout.addWidget(lbl_logo)

        self.nav_btns = {}
        for key, text, icon in [("results", "分配结果", "📊"), ("config", "全局设置", "⚙️"), ("groups", "分组管理", "👥")]:
            btn = QPushButton(f" {icon}  {text}")
            btn.setObjectName("SidebarBtn")
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda checked, k=key: self.switch_to(k))
            sidebar_layout.addWidget(btn)
            self.nav_btns[key] = btn

        lbl_hist_title = QLabel("📜 最近历史")
        lbl_hist_title.setStyleSheet(f"color: {COLORS['text_dim']}; font-size: 12px; padding: 40px 0 10px 25px;")
        sidebar_layout.addWidget(lbl_hist_title)

        self.history_frame = QWidget()
        self.history_layout = QVBoxLayout(self.history_frame)
        self.history_layout.setContentsMargins(15, 0, 15, 0)
        self.history_layout.setSpacing(5)
        sidebar_layout.addWidget(self.history_frame)
        sidebar_layout.addStretch()

        main_layout.addWidget(self.sidebar)

        self.main_container = QWidget()
        container_layout = QVBoxLayout(self.main_container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        self.stacked_widget = QStackedWidget()
        self.pages = {
            "results": ResultsPage(self),
            "config": ConfigPage(self),
            "groups": GroupsPage(self)
        }
        for page in self.pages.values():
            self.stacked_widget.addWidget(page)

        container_layout.addWidget(self.stacked_widget)

        self.statusBar = QStatusBar()
        self.statusBar.setSizeGripEnabled(False)
        self.statusBar.setStyleSheet(f"color: {COLORS['text_dim']}; padding: 6px 30px;")
        self.status_lbl = QLabel("就绪")
        self.statusBar.addWidget(self.status_lbl)
        container_layout.addWidget(self.statusBar)

        main_layout.addWidget(self.main_container)

    def set_status(self, text, is_error=False, duration=5000):
        color = COLORS['danger'] if is_error else COLORS['success'] if ("✅" in text or "保存成功" in text or "更新" in text or "更新为" in text or "导入" in text or "载入" in text) else COLORS['text_dim']
        self.status_lbl.setText(text)
        self.status_lbl.setStyleSheet(f"color: {color}; font-weight: bold; font-size: 12px;")

        if duration > 0:
            if self._status_timer:
                self.killTimer(self._status_timer)
            self._status_timer = self.startTimer(duration)

    def timerEvent(self, event):
        self.status_lbl.setText("就绪")
        self.status_lbl.setStyleSheet(f"color: {COLORS['text_dim']}; font-size: 12px;")
        self.killTimer(self._status_timer)
        self._status_timer = None

    def switch_to(self, page_key):
        if page_key != "results":
            self.pages["results"].commit_active_editor()

        for k, btn in self.nav_btns.items():
            btn.setChecked(k == page_key)
        
        self.stacked_widget.setCurrentWidget(self.pages[page_key])
        
        if page_key == "config": self.pages["config"].load_values()
        if page_key == "groups": self.pages["groups"].load_data()

    def update_history_sidebar(self):
        while self.history_layout.count():
            child = self.history_layout.takeAt(0)
            if child.widget(): child.widget().deleteLater()

        for item in self.model.history[:6]:
            ts = item['timestamp'].split(' ')[1]
            btn = QPushButton(f"🕐 {ts}")
            btn.setObjectName("SidebarBtn")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda checked=False, a=item['allocation'], t=ts: self.load_history_item(a, t))
            self.history_layout.addWidget(btn)

    def load_history_item(self, allocation, ts):
        self.switch_to("results")
        self.pages["results"].show_allocation(allocation)
        self.set_status(f"📅 已成功加载历史时间为【{ts}】的分配归档记录")

    def init_app_state(self):
        self.update_history_sidebar()
        if self.model.history:
            self.pages["results"].show_allocation(self.model.history[0]['allocation'])
        else:
            self.pages["results"].refresh_data()
        self.switch_to("results")


def main():
    app = QApplication(sys.argv)
    win = Application()
    win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()